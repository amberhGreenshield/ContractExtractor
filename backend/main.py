import os
import io
import json
import httpx
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from azure.storage.blob import BlobServiceClient

load_dotenv()

app = FastAPI(title="Contract Extractor API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Config ──────────────────────────────────────────────────────────────────
AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
BLOB_CONTAINER                  = os.getenv("BLOB_CONTAINER", "contracts")

AZURE_OPENAI_ENDPOINT   = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_KEY        = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4")

# ─── Blob Storage helpers ─────────────────────────────────────────────────────

def get_blob_client():
    return BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)

def list_blobs() -> list[dict]:
    client = get_blob_client()
    container = client.get_container_client(BLOB_CONTAINER)
    files = []
    for blob in container.list_blobs():
        name_lower = blob.name.lower()
        if name_lower.endswith(".pdf") or name_lower.endswith(".docx") or name_lower.endswith(".txt"):
            parts = blob.name.split("/")
            file_name = parts[-1]
            folder = "/".join(parts[:-1]) if len(parts) > 1 else "root"
            files.append({
                "id": blob.name,        # full blob path used as ID
                "name": file_name,
                "folder": folder,
                "path": blob.name,
                "size": blob.size,
            })
    # Sort by folder then filename so nested structure is obvious
    files.sort(key=lambda f: (f["folder"], f["name"]))
    return files

def download_blob_text(blob_name: str) -> str:
    client = get_blob_client()
    blob_name = blob_name.strip()
    print(f"DEBUG: Attempting to download blob: '{blob_name}'")
    print(f"DEBUG: Container: '{BLOB_CONTAINER}'")
    blob = client.get_blob_client(container=BLOB_CONTAINER, blob=blob_name)
    print(f"DEBUG: Blob URL: {blob.url}")
    content = blob.download_blob().readall()
    print(f"DEBUG: Downloaded {len(content)} bytes")

    name_lower = blob_name.lower().strip()

    if name_lower.endswith(".txt"):
        return content.decode("utf-8", errors="ignore")

    elif name_lower.endswith(".docx"):
        try:
            import docx
            doc = docx.Document(io.BytesIO(content))
            return "\n".join([p.text for p in doc.paragraphs])
        except Exception as e:
            return f"[Could not parse DOCX: {e}]"

    elif name_lower.endswith(".pdf"):
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages[:10]:
                    text_parts.append(page.extract_text() or "")
            return "\n".join(text_parts)
        except Exception as e:
            return f"[Could not parse PDF: {e}]"

    return "[Unsupported file type]"

# ─── Azure OpenAI extraction ──────────────────────────────────────────────────

EXTRACTION_PROMPT = """You are a contract analysis assistant. Your job is to extract four specific fields from a contract, regardless of its format or structure.

Contracts come in many formats — some have formal BETWEEN/AND blocks, some summarize parties in a preamble sentence like "Raj Rajbir of 2495577 ONTARIO INC.", some list parties in a table, and amendments may reference an original agreement. Read the entire document and use context clues.

FIELDS TO EXTRACT:

1. "supplier_name": The incorporated company or legal entity that is PROVIDING the services.
   - This is the Consultant / Supplier / Vendor / Contractor entity — the one being paid.
   - It is usually a numbered company (e.g. "2495577 ONTARIO INC.") or a named company.
   - It is NOT the client or buyer (the company paying for services, e.g. GreenShield, a bank, a hospital).
   - It may appear as: "Raj Rajbir of 2495577 ONTARIO INC." → supplier is "2495577 ONTARIO INC."
   - In amendments, it may refer back to the original agreement — still extract the supplier entity.

2. "contractor_name": The individual PERSON who is doing the work or who signed as the supplier representative.
   - Check the signature block first — look for the name signed on the supplier/consultant side.
   - Also check preamble sentences like "Raj Rajbir of 2495577 ONTARIO INC." → person is "Raj Rajbir".
   - This is always a human name, not a company name.
   - If the same person appears in both the preamble and the signature, that confirms it.

3. "hourly_rate": The per-hour fee charged by the consultant.
   - Search anywhere in the document: Compensation, Fees, Rate, Payment, Schedule sections.
   - Look for patterns like "$X per hour", "$X/hr", "hourly rate of $X".
   - If the rate is monthly or fixed (not hourly), return that value and note it in "notes".
   - Return as a clean string e.g. "$60/hr" or "$5,000/month".

4. "position_title": The job title or role the consultant is performing.
   - Look in Services, Scope of Work, Provision of Services, or Recitals sections.
   - It is often in quotes or bold e.g. "Senior Project Coordinator", "Business Analyst".
   - In amendments, it may replace a previous title — use the most recent one.

Return ONLY a valid JSON object with exactly these keys. No markdown, no explanation, nothing outside the JSON:
{
  "supplier_name": "...",
  "contractor_name": "...",
  "hourly_rate": "...",
  "position_title": "...",
  "notes": "note anything ambiguous, missing, or assumed here"
}

If a field genuinely cannot be found after reading the whole document, use null.

Contract text:
"""

async def extract_contract_info(text: str) -> dict:
    if not text.strip():
        return {"vendor_name": None, "contract_value": None, "contract_value_numeric": None, "currency": None, "notes": "Empty document"}

    truncated = text[:12000]
    url = f"{AZURE_OPENAI_ENDPOINT.rstrip('/')}/openai/deployments/{AZURE_OPENAI_DEPLOYMENT}/chat/completions?api-version=2024-08-01-preview"
    payload = {
        "messages": [
            {"role": "system", "content": "You extract structured information from contracts. Always respond with valid JSON only."},
            {"role": "user", "content": EXTRACTION_PROMPT + truncated}
        ],
        "temperature": 0,
        "max_tokens": 500,
    }

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            url,
            headers={"api-key": AZURE_OPENAI_KEY, "Content-Type": "application/json"},
            json=payload,
        )
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"].strip()

    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip("` \n")

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {"vendor_name": None, "contract_value": None, "contract_value_numeric": None, "currency": None, "notes": f"Parse error: {raw[:100]}"}

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"status": "ok"}

class TestRequest(BaseModel):
    text: str

@app.post("/contracts/test")
async def test_extraction(req: TestRequest):
    """Paste raw contract text and see what GPT-4 extracts. No blob storage needed."""
    try:
        result = await extract_contract_info(req.text)
        return {
            "extracted": result,
            "chars_processed": len(req.text),
            "truncated": len(req.text) > 12000,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/contracts/list")
async def list_contracts():
    try:
        files = list_blobs()
        return {"files": files, "total": len(files)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ExtractRequest(BaseModel):
    file_ids: list[str]

@app.post("/contracts/extract")
async def extract_contracts(req: ExtractRequest):
    results = []
    for blob_name in req.file_ids:
        try:
            text = download_blob_text(blob_name)
            extracted = await extract_contract_info(text)
            # derive a clean folder name from the blob path
            parts = blob_name.split("/")
            folder = "/".join(parts[:-1]) if len(parts) > 1 else "root"
            results.append({
                "file_id": blob_name,
                "file_name": parts[-1],
                "folder": folder,
                **extracted,
            })
        except Exception as e:
            results.append({
                "file_id": blob_name,
                "file_name": blob_name.split("/")[-1],
                "folder": "",
                "error": str(e),
            })
    return {"results": results}

@app.post("/contracts/export-excel")
async def export_excel(req: ExtractRequest):
    extraction = await extract_contracts(req)
    results = extraction["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Analysis"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1B3A5C")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="C0CCDA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", start_color="EEF3F8")
    currency_font = Font(name="Arial", size=10, bold=True, color="1B5E20")
    error_fill   = PatternFill("solid", start_color="FFEBEE")
    error_font   = Font(name="Arial", size=10, color="C62828")

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Contract Vendor & Cost Analysis"
    ws["A1"].font  = Font(name="Arial", bold=True, size=14, color="1B3A5C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Generated via Contract Extractor  •  {len(results)} contracts processed"
    ws["A2"].font  = Font(name="Arial", size=9, color="666666", italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    headers    = ["#", "File Name", "Folder", "Supplier (Company)", "Contractor (Person)", "Hourly Rate", "Position Title", "Notes"]
    col_widths = [5, 30, 25, 28, 22, 14, 28, 35]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 28

    for i, row in enumerate(results, 1):
        excel_row = i + 4
        is_alt    = i % 2 == 0
        has_error = "error" in row

        row_data = [
            i,
            row.get("file_name", ""),
            row.get("folder", ""),
            row.get("supplier_name") or ("ERROR" if has_error else "Not found"),
            row.get("contractor_name") or ("—" if not has_error else row.get("error", "")),
            row.get("hourly_rate") or "—",
            row.get("position_title") or "—",
            row.get("notes") or "—",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [2, 6, 7]))

            if has_error:
                cell.fill = error_fill
                cell.font = error_font
            else:
                if is_alt:
                    cell.fill = alt_fill
                cell.font = currency_font if col == 4 else Font(name="Arial", size=10)

        ws.row_dimensions[excel_row].height = 22

    summary_row = len(results) + 6
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws[f"A{summary_row}"].value = f"Total contracts analyzed: {len(results)}"
    ws[f"A{summary_row}"].font  = Font(name="Arial", bold=True, size=10, color="1B3A5C")
    ws[f"A{summary_row}"].alignment = Alignment(horizontal="left")

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contract_analysis.xlsx"},
    )
